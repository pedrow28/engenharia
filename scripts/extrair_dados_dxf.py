"""
extrair_dados_dxf.py
====================
Extrai dados de engenharia de arquivos DXF (convertidos de DWG)
e gera um relatório com os dados prontos para a planilha.

Uso:
    python extrair_dados_dxf.py <arquivo.dxf>
    python extrair_dados_dxf.py <pasta_com_dxf>

Dependências:
    pip install ezdxf
"""

import ezdxf
import sys
import os
import re
import json
import time
from collections import defaultdict
from pathlib import Path


# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

# Peso específico do concreto (kg/m³)
PESO_ESPECIFICO_CONCRETO = 2400

# Flag de debug — controlada por env var DEBUG_DXF=1.
# Quando desligada (padrão) suprime os prints [DEBUG ...] para manter o log limpo
# em execuções em lote (ex.: 260+ lajes).
DEBUG = os.environ.get('DEBUG_DXF', '').strip() not in ('', '0', 'false', 'False')


def dprint(*args, **kwargs):
    """print() condicional ao DEBUG. Sempre com flush=True quando ativado."""
    if DEBUG:
        kwargs.setdefault('flush', True)
        print(*args, **kwargs)


def iprint(*args, **kwargs):
    """print() de progresso — sempre visível, sempre com flush."""
    kwargs.setdefault('flush', True)
    print(*args, **kwargs)


# =============================================================================
# EXTRAÇÃO DO NOME DO ARQUIVO
# =============================================================================

def parse_filename(filename):
    """
    Extrai dados do nome do arquivo.
    Padrão: PEÇA-LARGURAxALTURAxCOMPRIMENTO-FORMA E ARMAÇÃO-RXX
    Exemplo: P5-30x80x1408-FORMA E ARMAÇÃO-R01

    Tolera espaços opcionais ao redor dos hífens separadores — alguns projetos
    de pilares nomeiam assim: "P10 - 40x50x2097-FORMA E ARMAÇÃO-R00".
    """
    name = Path(filename).stem
    # Remover sufixo de cópia do Windows ex: " (1)", " (2)"
    name = re.sub(r'\s*\(\d+\)$', '', name)

    # Padrão com seção: nome-LxAxC-tipo-revisao (com espaços opcionais nos hífens)
    match = re.match(
        r'^(.+?)\s*-\s*(\d+[xX]\d+(?:[xX][\d,.]+)?)\s*-\s*(.+?)\s*-\s*(R\d+)$',
        name,
    )
    if match:
        titulo = match.group(1).strip()
        secao_raw = match.group(2)
        parts = re.split(r'[xX]', secao_raw)
        if len(parts) == 3:
            largura, altura, comprimento = parts
            secao = f"{largura}x{altura}"
            comprimento = comprimento.replace(',', '.')
        else:
            secao = secao_raw
            comprimento = None
        return {
            'nome_arquivo': name,
            'titulo_peca': titulo,
            'secao': secao,
            'comprimento_cm': float(comprimento) if comprimento else None,
        }

    # Padrão sem seção (ex: lajes L201=L301-FORMA E ARMAÇÃO-R00)
    match2 = re.match(
        r'^(.+?)\s*-\s*(FORMA E ARMAÇÃO|FORMA|ARMAÇÃO)\s*-\s*(R\d+)$',
        name,
    )
    if match2:
        return {
            'nome_arquivo': name,
            'titulo_peca': match2.group(1).strip(),
            'secao': None,
            'comprimento_cm': None,
        }

    return {
        'nome_arquivo': name,
        'titulo_peca': name,
        'secao': None,
        'comprimento_cm': None,
    }


# =============================================================================
# EXTRAÇÃO DOS BLOCOS
# =============================================================================

def extract_notas(msp):
    """
    Extrai dados do bloco NOTAS.
    A=fck, B=Volume, C=Peso concreto, D=Peso peça, E=fcj (ou posição 5/7=G), I=Cobrimento

    O fcj pode estar em diferentes atributos dependendo do template:
    - Atributo 'E' (posição 5): padrão para pilares
    - Atributo 'G' (posição 7): encontrado nas vigas (20,3 MPa)
    Tenta cada um na ordem até encontrar valor válido.
    """
    for e in msp.query('INSERT'):
        if e.dxf.name == 'NOTAS':
            notas = {a.dxf.tag: a.dxf.text for a in e.attribs}
            def safe_float(key):
                val = notas.get(key, '-')
                if not val or str(val).strip() in ('-', '', '–'):
                    return None
                try:
                    return float(str(val).replace(',', '.'))
                except (ValueError, TypeError):
                    return None
            def extract_fcj():
                """
                Extrai fcj conforme o template de viga/pilar:
                - Tag H = nota 7 (LIBERAÇÃO DA PROTENSÃO fcj) — vigas protendidas
                - Tag E = nota 5 (LIBERAÇÃO P/ DESFORMA fcj)  — vigas sem protensão / pilares
                Prioridade: H > E. Tag G é a força de protensão (t/c) e NÃO é fcj.
                """
                # Protensão (nota 7) tem prioridade
                v = safe_float('H')
                if v is not None:
                    return v
                # Desforma (nota 5) — vigas sem protensão e pilares
                v = safe_float('E')
                if v is not None:
                    return v
                return None
            return {
                'fck_mpa': safe_float('A'),
                'volume_concreto_m3': safe_float('B'),
                'peso_concreto_kgf': safe_float('C'),
                'peso_peca_kgf': safe_float('D'),
                'fcj_mpa': extract_fcj(),
                'cobrimento_cm': safe_float('I'),
            }
    return {}


def extract_carimbo(msp):
    """
    Extrai dados do bloco CARIMBO.
    Atributos têm tag="X", usamos posição (índice) para identificar.
    """
    for e in msp.query('INSERT'):
        if e.dxf.name == 'CARIMBO':
            attribs = [a.dxf.text for a in e.attribs]
            if len(attribs) >= 16:
                nome_com_qtd = attribs[2]
                match = re.search(r'\((\d+)[xX]\)', nome_com_qtd)
                nome_limpo = re.sub(r'\(\d+[xX]\)', '', nome_com_qtd).strip()
                if match:
                    quantidade = int(match.group(1))
                elif '=' in nome_limpo:
                    quantidade = len(nome_limpo.split('='))
                else:
                    quantidade = None  # sem padrão explícito
                return {
                    'tipo_desenho': attribs[0],
                    'titulo_peca': nome_limpo,
                    'quantidade': quantidade,
                    'secao_comprimento': attribs[4],
                    'cliente': attribs[5],
                    'obra': attribs[6],
                    'data': attribs[12],
                    'folha': attribs[15],
                }
            break
    return {}


def extract_laje_carimbo(msp):
    """
    Extrai dados do bloco SM_formatoA4paraLajes (carimbo das lajes).
    [0]=tipo_desenho, [1]=titulo+qtd (ex: 'L-1001 (88x)'), [2]=seção+comprimento (ex: 'L-16x125x712,5')
    """
    for e in msp.query('INSERT'):
        if e.dxf.name == 'SM_formatoA4paraLajes':
            attribs = [a.dxf.text for a in e.attribs]
            if len(attribs) < 3:
                break

            nome_com_qtd = attribs[1]
            match_qtd = re.search(r'\((\d+)[xX]\)', nome_com_qtd)
            nome_limpo = re.sub(r'\s*\(\d+[xX]\)', '', nome_com_qtd).strip()
            if match_qtd:
                quantidade = int(match_qtd.group(1))
                quantidade_explicita = True
            elif '=' in nome_limpo:
                quantidade = len(nome_limpo.split('='))
                quantidade_explicita = False
            else:
                quantidade = None
                quantidade_explicita = False

            # Parse seção e comprimento: 'L-16x125x712,5' → secao='16x125', comp=712.5
            secao_raw = re.sub(r'^[A-Za-z]+-?', '', attribs[2])  # remove prefixo de letra
            parts = re.split(r'[xX]', secao_raw)
            if len(parts) >= 3:
                secao = f"{parts[0]}x{parts[1]}"
                comprimento_cm = float(parts[2].replace(',', '.'))
            elif len(parts) == 2:
                secao = f"{parts[0]}x{parts[1]}"
                comprimento_cm = None
            else:
                secao = secao_raw
                comprimento_cm = None

            return {
                'titulo_peca': nome_limpo,
                'quantidade': quantidade,
                'quantidade_explicita': quantidade_explicita,
                'secao': secao,
                'comprimento_cm': comprimento_cm,
            }
    return {}


def extract_laje_table(msp, doc):
    """
    Extrai dados das tabelas da laje via bloco ACAD_TABLE (*T10).
    Tabelas no bloco: RESUMO DE AÇO CP-190 RB, TENSÕES DO CONCRETO, DADOS PARA CONFERÊNCIA.
    Valores ficam na mesma linha (Y) do rótulo, em x≈317.
    """
    # Localizar o bloco de geometria do ACAD_TABLE
    table_block_name = None
    for e in msp:
        if e.dxftype() == 'ACAD_TABLE':
            try:
                table_block_name = e.dxf.geometry
            except Exception:
                pass
            break

    if not table_block_name or table_block_name not in doc.blocks:
        return {}

    blk = doc.blocks[table_block_name]

    def clean(txt):
        txt = re.sub(r'\{\\[^;]*;([^}]*)\}', r'\1', txt)
        txt = re.sub(r'\\[A-Za-z][^;]*;', '', txt).strip()
        return txt

    # Coletar textos agrupados por Y
    y_groups = defaultdict(list)
    for e in blk:
        if e.dxftype() == 'TEXT':
            txt = e.dxf.text.strip()
            if txt:
                y_groups[round(e.dxf.insert.y, 0)].append((e.dxf.insert.x, txt))
        elif e.dxftype() == 'MTEXT':
            txt = clean(e.text).strip()
            if txt and txt != '{':
                y_groups[round(e.dxf.insert.y, 0)].append((e.dxf.insert.x, txt))

    volume = fck = fcj = peso_protendido = None

    for items in y_groups.values():
        items_s = sorted(items, key=lambda i: i[0])

        for i, (x, txt) in enumerate(items_s):
            tu = txt.upper()

            if 'VOLUME DO CONCRETO' in tu:
                for _, v in items_s[i + 1:]:
                    try:
                        volume = float(v.replace(',', '.'))
                        break
                    except ValueError:
                        continue

            elif 'DE TRABALHO' in tu and 'FCK' in tu:
                for _, v in items_s[i + 1:]:
                    try:
                        fck = float(v.replace(',', '.'))
                        break
                    except ValueError:
                        continue

            elif ('PARA PROTENSÃO' in tu or 'PARA PROTENSAO' in tu) and 'FCJ' in tu:
                for _, v in items_s[i + 1:]:
                    try:
                        fcj = float(v.replace(',', '.'))
                        break
                    except ValueError:
                        continue

            # PESO TOTAL do RESUMO DE AÇO CP-190 RB: rótulo na coluna esquerda (x < 200)
            elif tu == 'PESO TOTAL' and x < 200:
                for _, v in items_s[i + 1:]:
                    try:
                        peso_protendido = float(v.replace(',', '.'))
                        break
                    except ValueError:
                        continue

    return {
        'volume_concreto_m3': volume,
        'fck_mpa': fck,
        'fcj_mpa': fcj,
        'peso_aco_protendido_kg': peso_protendido,
    }


# =============================================================================
# EXTRAÇÃO DAS TABELAS DE AÇO (LAYER ListaF)
# =============================================================================

def _extrair_peso_cp190rb(listaf, y_groups):
    """
    Detecta aço protendido CP190RB no layer ListaF.

    A tabela de lista de ferros tem dois tipos de linha com CP190RB:
    - Linha de item individual: "N° | CP190RB | diam | qtd | comp | peso"
      (o primeiro token da linha é o número do item — um inteiro)
    - Linha de resumo: "CP190RB | diam | comp_total | peso_metro | peso_total"
      (CP190RB é o primeiro token significativo da linha)

    O peso correto é o ÚLTIMO valor da linha de resumo (já inclui fator de
    alongamento). As linhas de item individual são ignoradas pois já estão
    contabilizadas no resumo.

    Retorna o peso total do aço protendido, ou 0 se não encontrado.
    """
    peso_protendido = 0.0
    encontrados = []

    for y, items in y_groups.items():
        items_sorted = sorted(items, key=lambda i: i[0])

        # Verificar se CP190RB está nesta linha
        cp_indices = [i for i, (x, txt) in enumerate(items_sorted)
                      if 'CP190RB' in txt.upper()]
        if not cp_indices:
            continue

        primeiro_cp_idx = cp_indices[0]

        # Determinar se é linha de item: verifica se há um inteiro imediatamente
        # antes do CP190RB (= número de item da lista de ferros)
        e_linha_item = False
        if primeiro_cp_idx > 0:
            txt_antes = items_sorted[primeiro_cp_idx - 1][1].strip()
            # Limpar formatação MTEXT
            txt_limpo = re.sub(r'\{\\[^;]*;([^}]*)\}', r'\1', txt_antes).strip()
            try:
                int(txt_limpo)
                e_linha_item = True
            except ValueError:
                pass

        if e_linha_item:
            # Linha de item individual: pular (já somada no resumo)
            continue

        # Linha de resumo: pegar o ÚLTIMO valor numérico (peso total com fator)
        valores = []
        for j in range(primeiro_cp_idx + 1, len(items_sorted)):
            txt_j = items_sorted[j][1].strip()
            try:
                val = float(txt_j.replace(',', '.'))
                if val > 0:
                    valores.append(val)
            except ValueError:
                continue

        if valores:
            peso_linha = valores[-1]
            encontrados.append({'y': y, 'x': items_sorted[primeiro_cp_idx][0], 'valor': peso_linha})
            peso_protendido += peso_linha

    if encontrados:
        dprint(f"  [DEBUG PROTENDIDO] CP190RB resumo encontrado ({len(encontrados)}):")
        for i, e in enumerate(encontrados):
            dprint(f"    [{i}] valor={e['valor']:.2f}  x={e['x']:.0f}  y={e['y']:.0f}")
    else:
        dprint(f"  [DEBUG PROTENDIDO] CP190RB não encontrado")

    return peso_protendido


def extract_peso_total_aco(msp):
    """
    Extrai valores PESO TOTAL (kg) das tabelas RESUMO DE AÇO.
    Retorna pesos separados: frouxo (armação principal), consolo e protendido (CP190RB).

    Lógica do protendido:
    - Se CP190RB for encontrado no ListaF, seu peso = protendido
    - O PESO TOTAL da tabela principal já inclui frouxo + protendido
    - Então: peso_frouxo = peso_total_tabela - peso_protendido
    """
    # Coletar textos do layer ListaF
    listaf = []
    for e in msp.query('TEXT'):
        if e.dxf.layer == 'ListaF':
            listaf.append((e.dxf.text, e.dxf.insert.x, e.dxf.insert.y))
    for e in msp.query('MTEXT'):
        if e.dxf.layer == 'ListaF':
            listaf.append((e.text, e.dxf.insert.x, e.dxf.insert.y))

    if not listaf:
        return {}

    # Agrupar por Y (linhas da tabela)
    y_groups = defaultdict(list)
    for txt, x, y in listaf:
        y_groups[round(y, 0)].append((x, txt))

    # --- Extrair peso do aço protendido (CP190RB) ---
    peso_protendido = _extrair_peso_cp190rb(listaf, y_groups)

    # --- Encontrar linhas "PESO TOTAL (kg)" e extrair valor numérico ---
    peso_entries = []
    for y, items in y_groups.items():
        items_sorted = sorted(items, key=lambda i: i[0])

        # Verificar se tem "PESO TOTAL" nesta linha
        peso_indices = [i for i, (x, txt) in enumerate(items_sorted) if 'PESO TOTAL' in txt.upper()]

        if peso_indices:
            for i in peso_indices:
                peso_x = items_sorted[i][0]

                # Procurar o valor numérico imediatamente DEPOIS do label
                if i + 1 < len(items_sorted):
                    prox_x, prox_txt = items_sorted[i+1]
                    try:
                        val = float(prox_txt.replace(',', '.'))
                        if prox_x - peso_x < 1000:
                            peso_entries.append({'y': y, 'x_label': peso_x, 'valor': val})
                    except ValueError:
                        continue

    if not peso_entries:
        return {}

    # DEBUG: todos os PESO TOTAL encontrados
    dprint(f"  [DEBUG CONSOLO] PESO TOTAL encontrados ({len(peso_entries)}):")
    if DEBUG:
        for i, pe in enumerate(peso_entries):
            dprint(f"    [{i}] valor={pe['valor']:.2f}  x={pe['x_label']:.0f}  y={pe['y']:.0f}")

    # --- Classificar tabelas: consolo vs armação principal ---
    # Cada bloco INSERT 'LISTA DE FERROS CONSOLO' tem atributo 'D' com multiplicidade
    # ex: '1x', '2x', '3x' — quando um consolo se repete na peça.
    consolo_block_info = []  # lista de (x, y, multiplicador)
    for e in msp.query('INSERT'):
        if e.dxf.name == 'LISTA DE FERROS CONSOLO':
            x = e.dxf.insert.x
            by = e.dxf.insert.y
            mult = 1
            for a in e.attribs:
                if a.dxf.tag == 'D':
                    m = re.match(r'(\d+)\s*[xX]', a.dxf.text.strip())
                    if m:
                        mult = int(m.group(1))
            consolo_block_info.append((x, by, mult))

    consolo_block_xs = [x for x, by, _ in consolo_block_info]
    dprint(f"  [DEBUG CONSOLO] Blocos INSERT 'LISTA DE FERROS CONSOLO': {len(consolo_block_info)} -> {[(f'x={x:.0f}', f'{m}x') for x, _, m in consolo_block_info]}")

    if consolo_block_xs:
        threshold = 1500

        def _dist_eucl(pe, bx, by):
            return ((pe['x_label'] - bx) ** 2 + (pe['y'] - by) ** 2) ** 0.5

        consolo = [e for e in peso_entries
                   if any(abs(e['x_label'] - bx) < threshold for bx in consolo_block_xs)]
        frouxo = [e for e in peso_entries if e not in consolo]
        dprint(f"  [DEBUG] Threshold={threshold}: consolo_entries={len(consolo)}, frouxo_entries={len(frouxo)}")

        # Fallback: quando todas as entradas foram para consolo (tabelas muito próximas em X),
        # usar distância euclidiana — a mais distante de qualquer bloco consolo = tabela principal.
        if not frouxo and len(consolo) > 1:
            dprint(f"  [DEBUG] Frouxo vazio — usando distância euclidiana para separar tabela principal.")
            def _min_dist(pe):
                return min(_dist_eucl(pe, bx, by) for bx, by, _ in consolo_block_info)
            sorted_by_dist = sorted(consolo, key=_min_dist, reverse=True)
            frouxo = [sorted_by_dist[0]]
            consolo = sorted_by_dist[1:]
            dprint(f"  [DEBUG] Após fallback euclidiano: frouxo={[e['valor'] for e in frouxo]}, consolo={[e['valor'] for e in consolo]}")

    elif len(peso_entries) == 1:
        frouxo = peso_entries
        consolo = []
    else:
        sorted_entries = sorted(peso_entries, key=lambda e: e['valor'], reverse=True)
        frouxo = [sorted_entries[0]]
        consolo = sorted_entries[1:]

    # Peso total da tabela principal (frouxo + protendido juntos)
    peso_total_tabela = max(e['valor'] for e in frouxo) if frouxo else None

    # Se há protendido, subtrair do total para obter o frouxo puro
    if peso_total_tabela is not None and peso_protendido > 0:
        peso_frouxo = peso_total_tabela - peso_protendido
        dprint(f"  [DEBUG PROTENDIDO] peso_total_tabela={peso_total_tabela:.2f} - protendido={peso_protendido:.2f} = frouxo={peso_frouxo:.2f}")
    else:
        peso_frouxo = peso_total_tabela

    # Consolo: somar com multiplicidade (atributo 'D' do bloco INSERT)
    # Para cada peso_entry do consolo, encontrar o bloco mais próximo e aplicar multiplicador
    def _mult_consolo(pe_x):
        candidatos = [(abs(pe_x - bx), mult) for bx, by, mult in consolo_block_info
                      if abs(pe_x - bx) < threshold]
        if candidatos:
            return min(candidatos, key=lambda t: t[0])[1]
        return 1

    n_consolos = len(consolo_block_info)
    if consolo:
        peso_consolo = sum(e['valor'] * _mult_consolo(e['x_label']) for e in consolo)
    else:
        peso_consolo = None

    dprint(f"  [DEBUG CONSOLO] frouxo={[e['valor'] for e in frouxo]}  consolo={[(e['valor'], _mult_consolo(e['x_label'])) for e in consolo]}")
    dprint(f"  [DEBUG CONSOLO] n_blocos={n_consolos} soma_consolo={peso_consolo}")

    return {
        'peso_aco_frouxo_kg': peso_frouxo,
        'peso_aco_consolo_kg': peso_consolo,
        'peso_aco_protendido_kg': peso_protendido,
        '_n_consolos': n_consolos,
    }


def extract_lista_ferros_detalhada(msp):
    """
    Extrai todos os itens da LISTA DE FERROS (layer ListaF) organizados por tabela.
    Retorna as linhas agrupadas por Y para referência.
    """
    listaf = []
    for e in msp.query('TEXT'):
        if e.dxf.layer == 'ListaF':
            listaf.append((e.dxf.text, round(e.dxf.insert.x, 1), round(e.dxf.insert.y, 1)))
    for e in msp.query('MTEXT'):
        if e.dxf.layer == 'ListaF':
            txt = re.sub(r'\{\\[^;]*;([^}]*)\}', r'\1', e.text)  # limpar formatação
            listaf.append((txt, round(e.dxf.insert.x, 1), round(e.dxf.insert.y, 1)))

    y_groups = defaultdict(list)
    for txt, x, y in listaf:
        y_groups[y].append((x, txt))

    rows = []
    for y in sorted(y_groups.keys(), reverse=True):
        items = sorted(y_groups[y], key=lambda i: i[0])
        row_text = ' | '.join(txt for _, txt in items)
        rows.append({'y': y, 'conteudo': row_text})
    return rows


# =============================================================================
# EXTRAÇÃO PRINCIPAL
# =============================================================================

def extrair_dados_completos(filepath):
    """Extrai todos os dados de um arquivo DXF."""
    try:
        doc = ezdxf.readfile(filepath)
    except IOError:
        iprint(f"  ERRO: Não foi possível ler: {filepath}")
        return None
    except ezdxf.DXFStructureError:
        iprint(f"  ERRO: DXF inválido: {filepath}")
        return None

    msp = doc.modelspace()
    filename = os.path.basename(filepath)

    dados_nome = parse_filename(filename)

    # Detectar tipo pelo nome do arquivo para escolher funções de extração corretas
    titulo_nome = dados_nome.get('titulo_peca', '')
    if re.match(r'^L', titulo_nome, re.IGNORECASE):
        _tipo_inicial = 'LAJE'
    elif re.match(r'^(VG|VL|VR|VTI|VTA|V\d)', titulo_nome):
        _tipo_inicial = 'VIGA'
    elif re.match(r'^P', titulo_nome):
        _tipo_inicial = 'PILAR'
    else:
        _tipo_inicial = 'OUTRO'

    if _tipo_inicial == 'LAJE':
        dados_carimbo = extract_laje_carimbo(msp)
        _laje_table = extract_laje_table(msp, doc)
        dados_notas = _laje_table  # mesmas chaves: volume_concreto_m3, fck_mpa, fcj_mpa
        dados_aco = {
            'peso_aco_frouxo_kg': 0,
            'peso_aco_consolo_kg': None,
            'peso_aco_protendido_kg': _laje_table.get('peso_aco_protendido_kg', 0),
            '_n_consolos': 0,
        }
    else:
        dados_carimbo = extract_carimbo(msp)
        dados_notas = extract_notas(msp)
        dados_aco = extract_peso_total_aco(msp)

    # Determinar tipo de peça (confirmação via carimbo)
    titulo = dados_carimbo.get('titulo_peca', dados_nome.get('titulo_peca', ''))
    if re.match(r'^(VG|VL|VR|VTI|VTA|V\d)', titulo):
        tipo_peca = 'VIGA'
    elif re.match(r'^L', titulo):
        tipo_peca = 'LAJE'
    elif re.match(r'^P', titulo):
        tipo_peca = 'PILAR'
    else:
        tipo_peca = 'OUTRO'

    # Valores base — peças igualadas (P11=P12 → 2) sempre têm prioridade
    # Para lajes igualadas com (Nx) explícito no carimbo (ex: L-201(75x)=L-301(75x)),
    # a quantidade total é Nx × número de igualadas (75 × 2 = 150).
    titulo_arquivo = dados_nome.get('titulo_peca', '')
    if '=' in titulo:
        n_iguais = len(titulo.split('='))
        carimbo_qtd = dados_carimbo.get('quantidade')
        if dados_carimbo.get('quantidade_explicita') and carimbo_qtd:
            quantidade = carimbo_qtd * n_iguais
        else:
            quantidade = n_iguais
    elif '=' in titulo_arquivo:
        quantidade = len(titulo_arquivo.split('='))
    elif dados_carimbo.get('quantidade') is not None:
        quantidade = dados_carimbo['quantidade']
    else:
        quantidade = 1

    if tipo_peca == 'LAJE':
        secao = dados_carimbo.get('secao') or dados_nome.get('secao', '')
        comprimento = dados_carimbo.get('comprimento_cm') or dados_nome.get('comprimento_cm')
    else:
        secao = dados_nome.get('secao') or dados_carimbo.get('secao_comprimento', '')
        comprimento = dados_nome.get('comprimento_cm')
    volume = dados_notas.get('volume_concreto_m3')
    fck = dados_notas.get('fck_mpa')
    fcj = dados_notas.get('fcj_mpa')

    # Pesos de aço
    peso_frouxo = dados_aco.get('peso_aco_frouxo_kg')
    peso_protendido = dados_aco.get('peso_aco_protendido_kg', 0)
    peso_consolo = dados_aco.get('peso_aco_consolo_kg')

    # Pilares nunca têm aço protendido — qualquer CP190RB detectado
    # (ex: grampos de ancoragem no consolo) é revertido para o frouxo.
    if tipo_peca == 'PILAR' and peso_protendido:
        dprint(f"  [PILAR] Revertendo CP190RB ({peso_protendido:.2f} kg) para frouxo — pilares não têm protendido.")
        peso_frouxo = (peso_frouxo or 0) + peso_protendido
        peso_protendido = 0

    # Peso total unitário = peso concreto + frouxo + protendido + consolo
    peso_concreto = volume * PESO_ESPECIFICO_CONCRETO if volume else None
    if peso_concreto is not None and peso_frouxo is not None:
        peso_total_unit = peso_concreto + peso_frouxo + (peso_protendido or 0) + (peso_consolo or 0)
    else:
        peso_total_unit = None

    # Taxas (kg/m³) — taxa frouxo inclui consolo (aço passivo total)
    aco_passivo = (peso_frouxo or 0) + (peso_consolo or 0)
    taxa_frouxo = aco_passivo / volume if (aco_passivo and volume) else None
    taxa_protendido = peso_protendido / volume if (peso_protendido and volume) else None

    # Totais (multiplicados pela quantidade)
    volume_total = volume * quantidade if volume else None
    peso_total_t = peso_total_unit * quantidade / 1000 if peso_total_unit else None
    peso_frouxo_t = peso_frouxo * quantidade / 1000 if peso_frouxo else None
    peso_protendido_t = peso_protendido * quantidade / 1000 if peso_protendido else None

    return {
        'A_tipo_peca': tipo_peca,
        'B_titulo_peca': titulo,
        'C_nome_desenho': dados_nome['nome_arquivo'],
        'D_quantidade': quantidade,
        'E_secao': secao,
        'F_comprimento_cm': comprimento,
        'G_volume_concreto_m3': volume,
        'H_volume_concreto_total_m3': volume_total,
        'I_fck_mpa': fck,
        'J_fcj_mpa': fcj,
        'K_peso_aco_frouxo_kg': peso_frouxo,
        'L_peso_aco_protendido_kg': peso_protendido,
        'M_peso_aco_consolo_kg': peso_consolo,
        'M_n_consolos': dados_aco.get('_n_consolos', 0),
        'N_peso_total_unitario_kg': peso_total_unit,
        'O_taxa_aco_frouxo_kg_m3': taxa_frouxo,
        'P_taxa_aco_protendido_kg_m3': taxa_protendido,
        'Q_peso_total_t': peso_total_t,
        'R_peso_total_aco_frouxo_t': peso_frouxo_t,
        'S_peso_total_aco_protendido_t': peso_protendido_t,
    }


# =============================================================================
# RELATÓRIO
# =============================================================================

def formatar_relatorio(dados):
    """Imprime relatório comparativo para validação."""
    if not dados:
        return "Nenhum dado extraído."

    campos = [
        ('A', 'Tipo de Peça', 'A_tipo_peca', ''),
        ('B', 'Título da Peça', 'B_titulo_peca', ''),
        ('C', 'Nome Desenho', 'C_nome_desenho', ''),
        ('D', 'Quantidade', 'D_quantidade', ''),
        ('E', 'Seção', 'E_secao', ''),
        ('F', 'Comprimento', 'F_comprimento_cm', 'cm'),
        ('G', 'Volume Concreto', 'G_volume_concreto_m3', 'm³'),
        ('H', 'Vol. Concreto Total', 'H_volume_concreto_total_m3', 'm³'),
        ('I', 'fck', 'I_fck_mpa', 'MPa'),
        ('J', 'fcj', 'J_fcj_mpa', 'MPa'),
        ('K', 'Peso Aço Frouxo', 'K_peso_aco_frouxo_kg', 'kg'),
        ('L', 'Peso Aço Protendido', 'L_peso_aco_protendido_kg', 'kg'),
        ('M', 'Peso Aço Consolo', 'M_peso_aco_consolo_kg', 'kg'),
        ('N', 'Peso Total Unitário', 'N_peso_total_unitario_kg', 'kg'),
        ('O', 'Taxa Aço Frouxo', 'O_taxa_aco_frouxo_kg_m3', 'kg/m³'),
        ('P', 'Taxa Aço Protendido', 'P_taxa_aco_protendido_kg_m3', 'kg/m³'),
        ('Q', 'Peso Total', 'Q_peso_total_t', 't'),
        ('R', 'Peso Total Frouxo', 'R_peso_total_aco_frouxo_t', 't'),
        ('S', 'Peso Total Protendido', 'S_peso_total_aco_protendido_t', 't'),
    ]

    lines = []
    lines.append("=" * 65)
    lines.append(f"  PEÇA: {dados['B_titulo_peca']}  |  {dados['C_nome_desenho']}")
    lines.append("=" * 65)
    for col, nome, chave, un in campos:
        val = dados.get(chave)
        if val is None:
            val_str = '-'
        elif isinstance(val, float):
            val_str = f"{val:,.4f}" if abs(val) < 1 else f"{val:,.2f}"
        else:
            val_str = str(val)
        extra = ''
        if chave == 'M_peso_aco_consolo_kg':
            n = dados.get('M_n_consolos', 0)
            if n > 0:
                extra = f'  ({n}x)'
        lines.append(f"  {col:2s} | {nome:25s} | {val_str} {un}{extra}")
    lines.append("")
    return "\n".join(lines)


# =============================================================================
# ATUALIZAÇÃO DA PLANILHA (Excel)
# =============================================================================

# Mapa tipo → nome da aba na planilha
_SHEET_MAP = {'PILAR': 'Pilares', 'VIGA': 'Vigas', 'LAJE': 'Lajes'}

# Mapa coluna (1-based) → chave do dict de dados extraídos
_MAPA_COLUNAS = {
    3: 'C_nome_desenho',             # C
    4: 'D_quantidade',               # D
    5: 'E_secao',                    # E
    6: 'F_comprimento_cm',           # F
    7: 'G_volume_concreto_m3',       # G
    9: 'I_fck_mpa',                  # I
    10: 'J_fcj_mpa',                 # J
    11: 'K_peso_aco_frouxo_kg',      # K
    12: 'L_peso_aco_protendido_kg',  # L
    13: 'M_peso_aco_consolo_kg',     # M
}


def _abrir_e_indexar(caminho_excel):
    """Abre o .xlsm e constrói o dict {titulo → linha} por aba.

    Retorna (wb, indices, proxima_livre). Em caso de erro, retorna (None, None, None).
    """
    try:
        import openpyxl
    except ImportError:
        iprint("ERRO: Biblioteca openpyxl não instalada. Instale com: pip install openpyxl")
        return None, None, None

    iprint(f"\nAbrindo planilha: {caminho_excel}")
    t0 = time.perf_counter()
    try:
        wb = openpyxl.load_workbook(caminho_excel, keep_vba=True)
    except Exception as e:
        iprint(f"ERRO ao abrir planilha: {e}")
        return None, None, None
    iprint(f"  [fase] Planilha carregada em {time.perf_counter() - t0:.1f}s")

    t_idx = time.perf_counter()
    indices: dict = {}         # {sheet_name: {titulo_norm: linha}}
    proxima_livre: dict = {}   # {sheet_name: próxima linha disponível}

    for sheet_name in set(_SHEET_MAP.values()):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        idx: dict = {}
        ultima_com_dado = 3  # cabeçalho até linha 3; dados começam em 4
        # iter_rows é MUITO mais rápido que ws.cell() em loop
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row,
                                min_col=2, max_col=2, values_only=False):
            cell = row[0]
            val = cell.value
            if val is not None:
                key = str(val).strip()
                if key and key not in idx:
                    idx[key] = cell.row
                ultima_com_dado = cell.row
        indices[sheet_name] = idx
        proxima_livre[sheet_name] = ultima_com_dado + 1
        iprint(f"  [fase] Índice '{sheet_name}': {len(idx)} peças existentes "
               f"(próxima linha livre: {proxima_livre[sheet_name]})")
    iprint(f"  [fase] Indexação concluída em {time.perf_counter() - t_idx:.1f}s")

    return wb, indices, proxima_livre


def _escrever_dados(wb, dados_lista, indices, proxima_livre, heartbeat_every=None):
    """Grava a lista de peças no workbook em memória usando lookup O(1).

    `indices` e `proxima_livre` são atualizados in-place quando novas linhas são criadas,
    permitindo que chamadas sucessivas (um lote após outro) permaneçam consistentes.

    Retorna dict com contadores: {'alteracoes', 'criadas', 'puladas', 'linhas': {titulo: linha}}
    """
    alteracoes = 0
    criadas = 0
    puladas = 0
    linhas_escritas: dict = {}  # {titulo_norm: (sheet_name, row)} — para verificação
    total = len(dados_lista)

    if heartbeat_every is None:
        heartbeat_every = max(1, total // 10) if total >= 20 else max(1, total)

    for i, dados in enumerate(dados_lista, 1):
        tipo = dados['A_tipo_peca']
        titulo = dados['B_titulo_peca']

        sheet_name = _SHEET_MAP.get(tipo)
        if not sheet_name:
            iprint(f"  [skip] {titulo}: tipo '{tipo}' sem aba correspondente.")
            puladas += 1
            continue
        if sheet_name not in wb.sheetnames:
            iprint(f"  [skip] {titulo}: aba '{sheet_name}' não existe.")
            puladas += 1
            continue

        ws = wb[sheet_name]
        idx = indices[sheet_name]
        key = str(titulo).strip()

        r = idx.get(key)
        if r is None:
            r = proxima_livre[sheet_name]
            proxima_livre[sheet_name] = r + 1
            idx[key] = r
            ws.cell(row=r, column=1).value = tipo     # A: Tipo
            ws.cell(row=r, column=2).value = titulo   # B: Título
            criadas += 1

        # Dados brutos
        for col_idx, chave in _MAPA_COLUNAS.items():
            valor = dados.get(chave)
            if valor is not None:
                ws.cell(row=r, column=col_idx).value = valor

        # Fórmulas
        ws.cell(row=r, column=8).value  = f'=G{r}*D{r}'                  # H: Volume total
        ws.cell(row=r, column=14).value = f'=G{r}*2400+K{r}+L{r}+M{r}'   # N: Peso unitário
        ws.cell(row=r, column=15).value = f'=(K{r}+M{r})/G{r}'           # O: Taxa frouxo
        ws.cell(row=r, column=16).value = f'=L{r}/G{r}'                  # P: Taxa protendido
        ws.cell(row=r, column=17).value = f'=N{r}*D{r}/1000'             # Q: Peso total (t)
        ws.cell(row=r, column=18).value = f'=K{r}*D{r}/1000'             # R: Total frouxo (t)
        ws.cell(row=r, column=19).value = f'=L{r}*D{r}/1000'             # S: Total protendido (t)

        alteracoes += 1
        linhas_escritas[key] = (sheet_name, r)

        if i % heartbeat_every == 0 or i == total:
            iprint(f"  [grav] {i}/{total} peças gravadas...")

    return {
        'alteracoes': alteracoes,
        'criadas': criadas,
        'puladas': puladas,
        'linhas': linhas_escritas,
    }


def _verificar_em_memoria(wb, dados_lista, indices):
    """Confirma que cada peça está na linha esperada do workbook em memória.

    Usa o dict de índices — muito barato. Não valida o save, apenas a escrita.
    Retorna (ok, lista_problemas).
    """
    problemas = []
    for dados in dados_lista:
        tipo = dados['A_tipo_peca']
        titulo = dados['B_titulo_peca']
        sheet_name = _SHEET_MAP.get(tipo)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue  # já foi reportado como "puladas" na escrita
        idx = indices.get(sheet_name, {})
        key = str(titulo).strip()
        r = idx.get(key)
        if r is None:
            problemas.append(f"{titulo} [sem linha no índice]")
            continue
        ws = wb[sheet_name]
        cell_val = ws.cell(row=r, column=2).value
        if cell_val is None or str(cell_val).strip() != key:
            problemas.append(f"{titulo} [linha {r} contém '{cell_val}']")
            continue
        # Sanity extra: a coluna C (nome_desenho) deve ter o mesmo valor que gravamos
        esperado_c = dados.get('C_nome_desenho')
        if esperado_c is not None:
            cell_c = ws.cell(row=r, column=3).value
            if cell_c != esperado_c:
                problemas.append(f"{titulo} [col C esperava '{esperado_c}', tem '{cell_c}']")
    return (len(problemas) == 0, problemas)


def _salvar(wb, caminho_excel):
    """Executa wb.save com tratamento de erro. Retorna (ok, mensagem_ou_None, duracao_s)."""
    t_sv = time.perf_counter()
    try:
        wb.save(caminho_excel)
        return True, None, time.perf_counter() - t_sv
    except PermissionError:
        return False, "arquivo aberto em outro programa — feche o Excel e rode de novo", time.perf_counter() - t_sv
    except Exception as e:
        return False, str(e), time.perf_counter() - t_sv


def _escrever_checkpoint_json(out_path, dados_acumulados):
    """Grava dados_acumulados num JSON de checkpoint. Silencioso em erro."""
    try:
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(dados_acumulados, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        iprint(f"  ⚠ Falha ao escrever checkpoint JSON: {e}")
        return False


def atualizar_planilha(caminho_excel, dados_extraidos):
    """
    Atualiza a planilha Excel com os dados extraídos em uma única passada.

    Thin wrapper sobre os helpers. Para lotes grandes (> BATCH_SIZE), o
    dispatcher em main() usa `_processar_em_lotes` ao invés desta função.
    """
    wb, indices, proxima_livre = _abrir_e_indexar(caminho_excel)
    if wb is None:
        return

    iprint(f"  {len(dados_extraidos)} peça(s) para gravar.")

    t_wr = time.perf_counter()
    stats = _escrever_dados(wb, dados_extraidos, indices, proxima_livre)
    iprint(f"  [fase] Escrita concluída em {time.perf_counter() - t_wr:.1f}s "
           f"({stats['alteracoes']} atualizadas, {stats['criadas']} criadas, "
           f"{stats['puladas']} puladas)")

    # Verificação em memória (barata — só confirma que escrevemos na linha certa)
    ok, problemas = _verificar_em_memoria(wb, dados_extraidos, indices)
    if not ok:
        iprint(f"  ⚠ verificação em memória: {len(problemas)} inconsistência(s)")
        for p in problemas[:5]:
            iprint(f"     - {p}")
    else:
        iprint(f"  ✓ verificação em memória: {stats['alteracoes']} peças ok")

    if stats['alteracoes'] > 0:
        iprint(f"  [fase] Salvando planilha... (pode demorar)")
        ok_save, err, dt_sv = _salvar(wb, caminho_excel)
        if ok_save:
            iprint(f"  [fase] Salvo em {dt_sv:.1f}s")
            iprint(f"\nPlanilha salva com sucesso! {stats['alteracoes']} peças atualizadas.")
        else:
            iprint(f"\nERRO ao salvar planilha: {err}")
            if 'aberto' in (err or ''):
                iprint("POR FAVOR, FECHE O ARQUIVO EXCEL E TENTE NOVAMENTE.")
    else:
        iprint("\nNenhuma alteração feita na planilha.")


# =============================================================================
# MAIN
# =============================================================================

def _resumo_linha(dados):
    """Linha-resumo compacta de uma peça (modo lote).
    Ex.: 'L-201  qtd=75  16x125x712,5  V=3.42m³  frouxo=180kg  prot=245kg'
    """
    titulo = dados.get('B_titulo_peca', '?')
    qtd = dados.get('D_quantidade')
    sec = dados.get('E_secao') or ''
    comp = dados.get('F_comprimento_cm')
    vol = dados.get('G_volume_concreto_m3')
    fr = dados.get('K_peso_aco_frouxo_kg')
    pr = dados.get('L_peso_aco_protendido_kg')
    co = dados.get('M_peso_aco_consolo_kg')

    partes = [f"{titulo}"]
    if qtd is not None:
        partes.append(f"qtd={qtd}")
    if sec:
        partes.append(f"{sec}" + (f"x{comp}" if comp else ""))
    if vol is not None:
        partes.append(f"V={vol:.2f}m³")
    if fr is not None:
        partes.append(f"frouxo={fr:.0f}kg")
    if pr:
        partes.append(f"prot={pr:.0f}kg")
    if co:
        partes.append(f"consolo={co:.0f}kg")
    return "  ".join(partes)


def _extrair_um_arquivo(arq, idx_global, total, modo_compacto):
    """Extrai dados de um arquivo DXF e imprime progresso. Retorna (dados|None, dt, erro|None)."""
    t_arq = time.perf_counter()
    if modo_compacto:
        iprint(f"[{idx_global}/{total}] {arq.name}")
    else:
        iprint(f"Processando: {arq.name}")

    try:
        dados = extrair_dados_completos(str(arq))
    except Exception as exc:
        dt = time.perf_counter() - t_arq
        iprint(f"  ✗ EXCEÇÃO em {arq.name}: {exc} ({dt:.1f}s)")
        return None, dt, str(exc)

    dt = time.perf_counter() - t_arq
    if dados:
        if modo_compacto:
            iprint(f"  ok ({dt:.1f}s)  {_resumo_linha(dados)}")
        else:
            iprint(formatar_relatorio(dados))
            iprint(f"  [tempo] {dt:.1f}s")
        return dados, dt, None

    iprint(f"  ✗ ERRO ao processar {arq.name} ({dt:.1f}s)")
    return None, dt, "sem dados"


def _processar_sequencial(arquivos, modo_compacto):
    """Extração pura (sem planilha) — usada quando não há caminho_excel.
    Retorna (todos, erros, t_acum_arq).
    """
    todos = []
    erros = []
    t_acum = 0.0
    total = len(arquivos)
    for i, arq in enumerate(arquivos, 1):
        dados, dt, _ = _extrair_um_arquivo(arq, i, total, modo_compacto)
        t_acum += dt
        if dados:
            todos.append(dados)
        else:
            erros.append(arq.name)
    return todos, erros, t_acum


def _resolver_dxf(dwg: Path):
    """Dado um .dwg, retorna o .dxf correspondente (existente) ou None.
    Tenta minúscula primeiro (saída padrão do ODA), depois maiúscula."""
    dxf_lower = dwg.with_suffix('.dxf')
    if dxf_lower.exists():
        return dxf_lower
    dxf_upper = dwg.with_suffix('.DXF')
    if dxf_upper.exists():
        return dxf_upper
    return None


def _processar_em_lotes(dwgs, pasta, caminho_excel, batch_size,
                        modo_compacto, checkpoint_json_path):
    """
    Pipeline COMPLETO em lotes: converter → extrair → gravar → verificar → salvar.
    Cada lote faz o processo inteiro antes de passar ao próximo.

    Ordem dentro de um lote:
    1. CONVERTER: DWGs pendentes do lote → DXFs (via ODA, staging hardlink).
       Pula DWGs que já têm DXF correspondente (resume automático).
    2. EXTRAIR: dados de cada DXF do lote (inclui os já pré-convertidos).
    3. GRAVAR: escreve os dados no workbook em memória.
    4. VERIFICAR: confirma em memória que cada peça está na linha certa.
    5. SALVAR: persiste o .xlsm em disco.
    6. CHECKPOINT: dump JSON incremental do acumulado.

    Resiliência: se qualquer fase quebrar, os lotes anteriores permanecem
    persistidos (no disco, no .xlsm e no checkpoint). Rerun pula os DWGs
    que já têm DXF E ainda regrava no Excel (idempotente via dict de índices).

    O workbook é aberto UMA ÚNICA VEZ antes do primeiro lote e mantido em
    memória até o fim — evita pagar o custo do load/indexação por lote.
    """
    # Import tardio para evitar ciclo e deixar erro claro se converter sumir
    try:
        from converter_dwg_dxf import encontrar_oda_converter, converter_lote
    except ImportError as e:
        iprint(f"ERRO ao importar converter_dwg_dxf: {e}")
        return [], [d.name for d in dwgs], 0.0

    oda_path = encontrar_oda_converter()
    if not oda_path:
        iprint("ERRO: ODA File Converter não encontrado!")
        iprint("Baixe em: https://www.opendesign.com/guestfiles/oda_file_converter")
        return [], [d.name for d in dwgs], 0.0
    iprint(f"ODA File Converter: {oda_path}")

    total = len(dwgs)
    n_lotes = (total + batch_size - 1) // batch_size
    ja_convertidos = sum(1 for d in dwgs if _resolver_dxf(d) is not None)

    iprint(f"\n{'═' * 56}")
    iprint(f"  PIPELINE EM LOTES: {total} arquivo(s) → {n_lotes} lote(s) de até {batch_size}")
    iprint(f"  Resume: {ja_convertidos} DWG(s) já com DXF — conversão será pulada")
    iprint(f"  Cada lote: converter → extrair → gravar → verificar → salvar")
    iprint(f"{'═' * 56}")

    # Abrir + indexar a planilha UMA VEZ — reaproveitado entre lotes.
    wb, indices, proxima_livre = _abrir_e_indexar(caminho_excel)
    if wb is None:
        iprint("  ⚠ Planilha indisponível — abortando pipeline.")
        return [], [d.name for d in dwgs], 0.0

    todos = []
    erros = []
    t_acum_arq = 0.0
    abortado = False

    for lote_idx in range(n_lotes):
        inicio = lote_idx * batch_size
        fim = min(inicio + batch_size, total)
        dwgs_lote = dwgs[inicio:fim]
        num_lote = lote_idx + 1

        iprint(f"\n{'━' * 56}")
        iprint(f"  LOTE {num_lote}/{n_lotes}  |  arquivos {inicio + 1}-{fim}/{total}")
        iprint(f"{'━' * 56}")

        t_lote = time.perf_counter()

        # ── Fase 1: CONVERTER DWGs pendentes do lote ─────────────────
        pendentes_lote = [d for d in dwgs_lote if _resolver_dxf(d) is None]
        reaproveitados = len(dwgs_lote) - len(pendentes_lote)

        if pendentes_lote:
            msg_reuse = f"  (+{reaproveitados} reaproveitado(s))" if reaproveitados else ""
            iprint(f"  [lote {num_lote}] convertendo {len(pendentes_lote)} DWG(s){msg_reuse}...")
            t_conv = time.perf_counter()
            try:
                movidos = converter_lote(oda_path, pasta, pendentes_lote,
                                         num_lote, n_lotes)
            except KeyboardInterrupt:
                iprint(f"  [lote {num_lote}] ⚠ interrompido pelo usuário durante conversão.")
                abortado = True
                break
            except Exception as exc:
                iprint(f"  [lote {num_lote}] ✗ EXCEÇÃO na conversão: {exc}")
                abortado = True
                break
            iprint(f"  [lote {num_lote}] conversão concluída em "
                   f"{time.perf_counter() - t_conv:.1f}s")
            if movidos == 0:
                iprint(f"  [lote {num_lote}] ⚠ nenhum DXF gerado — pulando lote.")
                erros.extend(d.name for d in pendentes_lote)
                continue
        else:
            iprint(f"  [lote {num_lote}] todos os {len(dwgs_lote)} DWG(s) já convertidos — pulando conversão.")

        # ── Fase 2: EXTRAIR DXFs do lote ─────────────────────────────
        iprint(f"  [lote {num_lote}] extraindo {len(dwgs_lote)} DXF(s)...")
        t_extr = time.perf_counter()
        dados_lote = []
        for i, dwg in enumerate(dwgs_lote, 1):
            dxf = _resolver_dxf(dwg)
            if dxf is None:
                iprint(f"  [lote {num_lote}] ⚠ DXF faltando para {dwg.name} — pulando.")
                erros.append(dwg.name + " (sem DXF)")
                continue
            global_i = inicio + i
            dados, dt, _ = _extrair_um_arquivo(dxf, global_i, total, modo_compacto)
            t_acum_arq += dt
            if dados:
                dados_lote.append(dados)
            else:
                erros.append(dxf.name)
        iprint(f"  [lote {num_lote}] extração: {len(dados_lote)}/{len(dwgs_lote)} ok "
               f"em {time.perf_counter() - t_extr:.1f}s")

        if not dados_lote:
            iprint(f"  [lote {num_lote}] ⚠ lote sem dados — pulando gravação.")
            continue

        # ── Fase 3: GRAVAR no workbook em memória ────────────────────
        iprint(f"  [lote {num_lote}] gravando no workbook...")
        t_wr = time.perf_counter()
        stats = _escrever_dados(wb, dados_lote, indices, proxima_livre,
                                heartbeat_every=max(10, len(dados_lote) // 3))
        iprint(f"  [lote {num_lote}] escrita: {stats['alteracoes']} alteradas, "
               f"{stats['criadas']} criadas, {stats['puladas']} puladas "
               f"em {time.perf_counter() - t_wr:.1f}s")

        # ── Fase 4: VERIFICAR em memória ─────────────────────────────
        ok_ver, problemas = _verificar_em_memoria(wb, dados_lote, indices)
        if not ok_ver:
            iprint(f"  [lote {num_lote}] ⚠ verificação: {len(problemas)} inconsistência(s)")
            for p in problemas[:5]:
                iprint(f"       - {p}")
            if len(problemas) > 5:
                iprint(f"       ... e mais {len(problemas) - 5}")
        else:
            iprint(f"  [lote {num_lote}] ✓ verificação: {stats['alteracoes']} peças confirmadas")

        # ── Fase 5: SALVAR .xlsm (persistir lote no disco) ───────────
        iprint(f"  [lote {num_lote}] salvando planilha...")
        ok_save, err, dt_sv = _salvar(wb, caminho_excel)
        if not ok_save:
            iprint(f"  [lote {num_lote}] ✗ SAVE FALHOU em {dt_sv:.1f}s: {err}")
            if 'aberto' in (err or ''):
                iprint("  ⚠ FECHE O EXCEL e rode novamente — os lotes anteriores já foram persistidos.")
            iprint(f"  ⚠ Abortando lotes restantes. Lotes {num_lote}..{n_lotes} NÃO foram salvos.")
            abortado = True
            break
        iprint(f"  [lote {num_lote}] ✓ salvo em {dt_sv:.1f}s")

        # Confirmado: lote persistido no .xlsm. Pode adicionar ao acumulado.
        todos.extend(dados_lote)

        # ── Fase 6: Checkpoint JSON incremental ──────────────────────
        if checkpoint_json_path is not None:
            _escrever_checkpoint_json(checkpoint_json_path, todos)

        iprint(f"  [lote {num_lote}] TOTAL: {time.perf_counter() - t_lote:.1f}s  "
               f"|  acumulado: {len(todos)}/{total} peças persistidas")

    if abortado:
        iprint(f"\n[resumo pipeline] ⚠ Interrompido — {len(todos)}/{total} peças persistidas antes do erro.")
    else:
        iprint(f"\n[resumo pipeline] ✓ {n_lotes} lote(s) processados.")

    return todos, erros, t_acum_arq


def _listar_dwgs_ordenado(pasta: Path):
    """Lista todos os DWGs da pasta (case-insensitive, ordenados pelo nome)."""
    vistos = {}
    for p in pasta.iterdir():
        if p.is_file() and p.suffix.lower() == '.dwg':
            vistos[p.name.lower()] = p
    return sorted(vistos.values(), key=lambda p: p.name.lower())


def main():
    if len(sys.argv) < 2:
        iprint("Uso: python extrair_dados_dxf.py <arquivo.dxf ou pasta> [caminho_planilha.xlsm]")
        iprint("  Pipeline completo: converte DWG→DXF, extrai dados e atualiza a planilha.")
        iprint("  Se a pasta só tiver DXFs (sem DWGs), apenas extrai — modo legado.")
        iprint("  Env:")
        iprint("    DEBUG_DXF=1    ativa prints de debug detalhados")
        iprint("    BATCH_SIZE=N   tamanho do lote (padrão 50; 0 desativa batching)")
        sys.exit(1)

    caminho = sys.argv[1]

    # Verificar se foi passado o excel como segundo argumento
    caminho_excel = None
    if len(sys.argv) > 2:
        caminho_excel = sys.argv[2]
        if not os.path.exists(caminho_excel):
            iprint(f"AVISO: Planilha não encontrada: {caminho_excel}")
            caminho_excel = None

    # Classificar entrada: arquivo único | pasta com DWGs (pipeline) | pasta só DXF (legado)
    entrada = Path(caminho)
    modo_pipeline = False
    dwgs: list = []
    arquivos: list = []

    if entrada.is_dir():
        dwgs = _listar_dwgs_ordenado(entrada)
        if dwgs and caminho_excel:
            modo_pipeline = True
            iprint(f"Encontrados {len(dwgs)} arquivo(s) DWG\n")
        else:
            arquivos = sorted(entrada.glob('*.dxf')) + sorted(entrada.glob('*.DXF'))
            # Dedup case-insensitive
            vistos = {p.name.lower(): p for p in arquivos}
            arquivos = sorted(vistos.values(), key=lambda p: p.name.lower())
            if not arquivos:
                iprint(f"Nenhum .dwg ou .dxf em: {caminho}")
                sys.exit(1)
            if dwgs and not caminho_excel:
                iprint(f"{len(dwgs)} DWG(s) encontrados mas sem planilha — extraindo apenas DXFs existentes.\n")
            else:
                iprint(f"Encontrados {len(arquivos)} arquivo(s) DXF (modo legado — sem conversão)\n")
    else:
        arquivos = [entrada]

    total = len(dwgs) if modo_pipeline else len(arquivos)

    # Modo compacto do log: com muitos arquivos, o relatório cheio polui demais.
    modo_compacto = total > 10
    if modo_compacto:
        iprint(f"Log compacto ativo ({total} arquivos). Use DEBUG_DXF=1 para detalhes.\n")

    # Batch size via env var (padrão 50; 0 desativa batching)
    try:
        batch_size = int(os.environ.get('BATCH_SIZE', '50'))
    except ValueError:
        batch_size = 50

    # Caminho do checkpoint JSON
    base = entrada.parent if entrada.is_file() else entrada
    checkpoint_path = base / 'dados_extraidos.json'

    t_geral = time.perf_counter()

    if modo_pipeline:
        # Pipeline completo em lotes: convert → extract → write → verify → save
        bs = batch_size if batch_size > 0 else total
        iprint(f"Pipeline em lotes: {bs} por lote (total={total}, excel='{Path(caminho_excel).name}')")
        todos, erros, t_acum_arq = _processar_em_lotes(
            dwgs, entrada, caminho_excel, bs, modo_compacto, checkpoint_path
        )
    else:
        # Modo legado: só extração (sem conversão). Usado quando a pasta já tem
        # DXFs prontos, ou quando foi passado um arquivo avulso, ou quando não
        # há planilha.
        motivo = ("arquivo único" if entrada.is_file()
                  else "sem planilha — só extração" if caminho_excel is None
                  else "pasta só com DXF")
        iprint(f"Modo legado ({motivo}) — sem conversão.")
        todos, erros, t_acum_arq = _processar_sequencial(arquivos, modo_compacto)

        # Legado: grava JSON e chama a planilha no final
        if todos:
            _escrever_checkpoint_json(checkpoint_path, todos)
            iprint(f"Dados salvos em: {checkpoint_path}")
            if caminho_excel:
                atualizar_planilha(caminho_excel, todos)

    # Resumo final
    dt_total = time.perf_counter() - t_geral
    media = (t_acum_arq / total) if total else 0.0
    iprint(f"\n[resumo extração] {len(todos)}/{total} ok  "
           f"| tempo total: {dt_total:.1f}s  "
           f"| média extração: {media:.2f}s/arq")
    if erros:
        iprint(f"[resumo extração] falhas: {len(erros)} -> {', '.join(erros[:5])}"
               + ("..." if len(erros) > 5 else ""))

    # Garante que o JSON final reflete o estado acumulado no modo pipeline
    # (nos lotes já está lá, mas não custa reescrever ao final).
    if todos and modo_pipeline:
        _escrever_checkpoint_json(checkpoint_path, todos)
        iprint(f"Dados finais em: {checkpoint_path}")

    iprint(f"[fim] tempo total do extractor: {dt_total:.1f}s")
    return todos


if __name__ == '__main__':
    main()
